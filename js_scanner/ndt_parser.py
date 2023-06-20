import openpyxl.worksheet.worksheet
from bs4 import BeautifulSoup, Tag
from openpyxl import load_workbook
import requests
import re
import time
import random


def read_excel(file_path: str) -> list:
    wb = load_workbook(file_path)
    sheet = wb.get_sheet_by_name('Sheet1')
    result = []
    for row in range(2, 405):
        try:
            ip = sheet[f'A{row}'].value
            tech = sheet[f'B{row}'].value
            level = sheet[f'C{row}'].value
            cvss = sheet[f'D{row}'].value
            desc = sheet[f'E{row}'].value
            cve_raw = sheet[f'F{row}'].value
            recom = sheet[f'G{row}'].value
            href = sheet[f'H{row}'].value
            port = sheet[f'I{row}'].value
            found_cve_raw = re.findall(r'CVE-\d{4}-\d{4,7}', cve_raw)
            result.append(
                {'ip': ip,
                 'tech': tech,
                 'level': level,
                 'cvss': cvss,
                 'desc': desc,
                 'cves': [cve for cve in found_cve_raw],
                 'recom': recom,
                 'href': href,
                 'port': port
                 }
            )
        except Exception as ex:
            print(ex)
            return result
    wb.close()
    return result


def parse_description_in_nvd_nist_gov(cve: str) -> str:
    is_request_done = False
    while not is_request_done:
        try:
            time.sleep(random.randint(1, 5))
            response = requests.get(f'https://nvd.nist.gov/vuln/detail/{cve}')
            if isinstance(response, requests.Response):
                soup = BeautifulSoup(response.text, 'lxml')
                div = soup.find('div', class_='col-lg-9 col-md-7 col-sm-12')
                description_div = div.find('p')
                if isinstance(description_div, Tag):
                    description_text = description_div.text
                    return description_text
                raise ValueError('I got wrong Tag(')
            raise ValueError('I got wrong response(')
        except Exception as ex:
            print('Request error', ex)
    raise ValueError


if __name__ == '__main__':
    str_91 = '91.245.41.0'
    str_31 = '31.41.241.0'
    str_144 = '185.117.144.0'
    str_145 = '185.117.145.0'
    wb = load_workbook('new.xlsx')
    wb.create_sheet(str_91)
    current_91_row = 0
    wb.create_sheet(str_31)
    current_31_row = 0
    wb.create_sheet(str_144)
    current_144_row = 0
    wb.create_sheet(str_145)
    current_145_row = 0
    parsed_all = read_excel('C:\\Users\\Салават\\Documents\\91.245.41.95.xlsx')
    index = 0
    for d in parsed_all:
        index += 1
        print('Index: ', index)
        for cve in d['cves']:
            description = parse_description_in_nvd_nist_gov(cve)
            ip = d['ip']
            subnet = re.search(r'\d{1,3}.\d{1,3}.\d{1,3}', ip).group(0)
            if subnet is not None:
                sheet:  openpyxl.worksheet.worksheet.Worksheet = wb.get_sheet_by_name(f'{subnet}.0')
                row = sheet.max_row + 1
                sheet[f'A{row}'].value = d['ip']
                sheet[f'B{row}'].value = d['tech']
                sheet[f'C{row}'].value = d['level']
                sheet[f'D{row}'].value = d['cvss']
                sheet[f'E{row}'].value = description
                sheet[f'F{row}'].value = cve
                sheet[f'G{row}'].value = d['recom']
                sheet[f'H{row}'].value = d['href']
                sheet[f'I{row}'].value = d['port']
                print(sheet[f'A{row}'].value)
    wb.save('salavat_cool.xlsx')
    wb.close()
